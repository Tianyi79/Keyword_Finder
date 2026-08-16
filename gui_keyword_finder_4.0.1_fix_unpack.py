"""GUI Keyword Finder 4.0 - Local Document Search Workbench

A local-first PySide6 desktop application for searching multiple documents,
reviewing structured hits, previewing positioned PDF matches, saving selected
quotes, and exporting results. Searches run in a background QThread so the UI
remains responsive.

Supported input paths
---------------------
PDF (.pdf)
    Opened directly with PyMuPDF. Every page is scanned. PDF words, line
    reconstruction data, and bounding boxes are used for exact-word results,
    context, and highlighted page previews.

CSV (.csv)
    Read directly with Python's csv module. The reader tries utf-8-sig,
    utf-8, gb18030, and latin-1 in that order.

Excel (.xlsx, .xlsm, .xltx, .xltm)
    Read directly with openpyxl in read-only, data-only mode. This searches
    stored cell values and cached formula results, not formula expressions.

Other non-PDF files (for example .docx, .pptx, or legacy formats)
    Extracted as text through Kreuzberg when Kreuzberg is installed. PDF,
    CSV, and supported Excel searches do not require Kreuzberg.

How searchable content is structured
------------------------------------
PDF
    Results store a 1-based page number and, when available, a PDF rectangle
    (x0, y0, x1, y1). Word lines are reconstructed from PyMuPDF's block and
    line identifiers to create snippets and nearby context.

CSV and Excel
    Each physical row becomes one searchable string. Empty cells are skipped
    and remaining values are joined with " | ". Results store a 1-based row
    number. Excel worksheet names and cell coordinates are not stored, so rows
    with the same number on different worksheets cannot be distinguished in
    exported results and may be grouped together in Page view.

Kreuzberg-extracted files
    Extracted content is split into lines. Results store a 1-based line number,
    the matching line as the snippet, and one neighboring line on each side as
    context.

Matching modes
--------------
Plain
    CSV, Excel, and extracted text produce one Hit for every accepted literal
    occurrence. PDFs use a cached word lookup for ASCII whole-word searches;
    other PDF literals and phrases use PyMuPDF page.search_for().

Regex
    CSV, Excel, and extracted text use Python regular expressions and produce
    one Hit per accepted match. PDFs run the expression against extracted page
    text, then try to locate the detected literal with page.search_for() for a
    best-effort highlight rectangle. PDF regex context is limited to the
    detected match text.

Fuzzy
    CSV, Excel, and Kreuzberg-extracted text compare the keyword with individual
    tokens using difflib.SequenceMatcher and the configured threshold. At most
    one best fuzzy token is emitted per keyword for each source row or line.
    PDF fuzzy mode currently follows the normal literal PDF search path; it does
    not apply token-similarity matching to PDF words.

Case and whole-word behavior
----------------------------
Case sensitivity is applied explicitly to CSV, Excel, extracted text, and the
cached PDF word lookup. PDF searches routed through page.search_for() use
PyMuPDF's own search semantics. Whole-word enforcement applies to ASCII
letter/digit/underscore keywords. When enabled, the detected ASCII token must
equal the keyword, so approximate fuzzy matches can be rejected even when their
similarity score passes the threshold.

Results and views
-----------------
Each Hit records:
    file path, file type, keyword, detected text, page/line/row location,
    optional PDF rectangle, snippet, and context.

Hit view
    Displays one table row for every Hit.

Page view
    Groups Hits by file and stored document location. For PDFs this is normally
    the page number; for extracted text it is the line number; for CSV/Excel it
    is the row number. Page view shows a representative Hit and the number of
    grouped matches.

PDF preview and navigation
--------------------------
The separate Preview window can:
    - render a PDF page and outline the selected Hit rectangle;
    - zoom with buttons or Ctrl+Wheel, fit width, or fit page;
    - show reconstructed context and full extracted page text;
    - move to the previous or next visible result;
    - open a PDF at the selected page in SumatraPDF on Windows when available,
      otherwise through the system viewer.

Non-PDF Hits show text in the Preview window but do not have a rendered page
image. A PDF regex Hit without a recovered rectangle also has no image highlight.

Clips and exports
-----------------
Text selected in the Preview context or page-text panel can be saved as a Clip.
A Clip stores its creation time, source path/type, page/line/row location,
keyword, detected text, selected text, and an optional note field.

Exports:
    - search results as CSV or XLSX;
    - clips as Markdown or CSV;
    - keywords as plain text;
    - a reusable AI keyword-generation prompt as plain text.

Caching and runtime behavior
----------------------------
PDF pages are scanned sequentially in the worker thread. Word lists, line
indexes, and case-specific exact-word lookups are built and cached when the
active search path needs them. Kreuzberg-extracted text is also cached. Direct
CSV and Excel row data is reread for every search and is not stored in DiskCache.

DiskCache stores gzip-compressed pickle files. Cache keys include the resolved
source path, integer modification time, file size, and a cache-purpose suffix.
The configured cache limit applies to cache entries, which are trimmed by most
recent modification time. The Preview window also maintains small in-memory LRU
caches for open PDF documents and rendered page images.

Keyword workflow
----------------
Keywords may be entered one per line or separated by commas. Empty values are
removed and duplicates are discarded according to the current case-sensitivity
setting. Keyword lists can be imported from text or from all non-empty CSV cells.

Configuration and local data
----------------------------
Normal mode stores config.json, cache files, and rotating logs under:
    ~/.keyword_finder/

Portable mode stores them beside the script under:
    .keyword_finder_data/

Default search settings are case-sensitive plain matching, ASCII whole-word
matching enabled, fuzzy threshold 0.86, Hit view, disk caching enabled, and PDF
preview zoom 2.0.

Install
-------
Core PDF/CSV/Excel functionality:
    python -m pip install PySide6 pymupdf pillow openpyxl

Optional extraction for other document formats:
    python -m pip install kreuzberg

Run
---
    python gui_keyword_finder_4.0.1_fix_unpack.py

Portable mode
-------------
    python gui_keyword_finder_4.0.1_fix_unpack.py --portable
"""

from __future__ import annotations

import csv
import gzip
import json
import logging
import os
import pickle
import platform
import re
import subprocess
import sys
from dataclasses import dataclass, asdict
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Any

import fitz  # PyMuPDF
from PIL import Image, ImageQt, ImageDraw
from openpyxl import Workbook, load_workbook

# Kreuzberg is optional: PDF, CSV, and supported Excel searches work without it.
try:
    from kreuzberg import batch_extract_files_sync, ExtractionConfig
    KREUZBERG_AVAILABLE = True
except Exception:
    batch_extract_files_sync = None  # type: ignore
    ExtractionConfig = None  # type: ignore
    KREUZBERG_AVAILABLE = False

from PySide6.QtCore import Qt, QThread, Signal, QTimer, QUrl, QObject, QEvent
from PySide6.QtGui import QPixmap, QDesktopServices, QAction, QKeySequence
from PySide6.QtWidgets import (
    QApplication,
    QMainWindow,
    QWidget,
    QFileDialog,
    QHBoxLayout,
    QVBoxLayout,
    QLabel,
    QTableWidget,
    QTableWidgetItem,
    QMessageBox,
    QHeaderView,
    QListWidget,
    QListWidgetItem,
    QTextEdit,
    QScrollArea,
    QGroupBox,
    QDialog,
    QDialogButtonBox,
    QPushButton,
    QAbstractItemView,
    QProgressBar,
    QComboBox,
    QCheckBox,
    QLineEdit,
    QFormLayout,
)


# -------------------- Data model --------------------


@dataclass
class Hit:
    file_path: str
    file_type: str  # "PDF" or "DOC/PPT/XLS/…"
    keyword: str
    detected_word: str
    page: Optional[int]  # PDF only (1-based)
    line_no: Optional[int]  # text-like non-PDF files (1-based)
    row_no: Optional[int]  # spreadsheet / CSV row number (1-based)
    rect: Optional[Tuple[float, float, float, float]]  # PDF bbox in page coords
    snippet: str
    context: str


@dataclass
class Clip: 
    created_at: str
    file_path: str
    file_type: str
    page: Optional[int]
    line_no: Optional[int]
    row_no: Optional[int]
    keyword: str
    detected_word: str
    selected_text: str
    note: str = ""


@dataclass
class AppConfig:
    case_sensitive: bool = True
    whole_word: bool = True
    search_mode: str = "plain"  # plain | regex | fuzzy
    fuzzy_threshold: float = 0.86
    external_viewer: str = "auto"  # auto | sumatra | default
    enable_cache: bool = True
    cache_max_files: int = 250
    last_dir: str = ""
    results_view: str = "hits"  # hits | pages
    zoom_default: float = 2.0


# -------------------- Paths / config / logging --------------------


def _is_windows() -> bool:
    return platform.system().lower().startswith("win")


class AppPaths:
    def __init__(self, portable: bool):
        self.portable = portable
        if portable:
            base = Path(__file__).resolve().parent / ".keyword_finder_data"
        else:
            base = Path.home() / ".keyword_finder"
        self.base = base
        self.config_path = base / "config.json"
        self.cache_dir = base / "cache"
        self.logs_dir = base / "logs"
        self.ensure()

    def ensure(self):
        self.base.mkdir(parents=True, exist_ok=True)
        self.cache_dir.mkdir(parents=True, exist_ok=True)
        self.logs_dir.mkdir(parents=True, exist_ok=True)


def setup_logging(logs_dir: Path) -> None:
    logs_dir.mkdir(parents=True, exist_ok=True)
    log_file = logs_dir / "app.log"
    logger = logging.getLogger()
    logger.setLevel(logging.INFO)
    fmt = logging.Formatter("%(asctime)s %(levelname)s %(message)s")

    # file handler (simple rotation by size)
    from logging.handlers import RotatingFileHandler

    fh = RotatingFileHandler(str(log_file), maxBytes=1_000_000, backupCount=3, encoding="utf-8")
    fh.setFormatter(fmt)
    logger.addHandler(fh)

    sh = logging.StreamHandler(sys.stdout)
    sh.setFormatter(fmt)
    logger.addHandler(sh)

    logging.info("Logging initialized: %s", log_file)


def load_config(path: Path) -> AppConfig:
    try:
        if path.exists():
            data = json.loads(path.read_text(encoding="utf-8"))
            cfg = AppConfig(**{k: data.get(k, getattr(AppConfig(), k)) for k in asdict(AppConfig()).keys()})
            return cfg
    except Exception as e:
        logging.warning("Failed to load config: %s", e)
    return AppConfig()


def save_config(path: Path, cfg: AppConfig) -> None:
    try:
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text(json.dumps(asdict(cfg), ensure_ascii=False, indent=2), encoding="utf-8")
    except Exception as e:
        logging.warning("Failed to save config: %s", e)


# -------------------- Disk cache --------------------


class DiskCache:
    """Tiny disk cache keyed by (path, mtime, size, extra).

    Stores gzip-compressed pickle blobs.
    """

    def __init__(self, cache_dir: Path, enable: bool = True, max_files: int = 250):
        self.cache_dir = cache_dir
        self.enable = enable
        self.max_files = max_files
        self.cache_dir.mkdir(parents=True, exist_ok=True)

    def _key(self, file_path: str, extra: str = "") -> str:
        p = Path(file_path)
        try:
            st = p.stat()
            mtime = int(st.st_mtime)
            size = int(st.st_size)
        except Exception:
            mtime, size = 0, 0
        safe = re.sub(r"[^A-Za-z0-9_.-]+", "_", str(p.resolve()))
        return f"{safe}__{mtime}_{size}__{extra}"[:240]

    def get(self, file_path: str, extra: str = "") -> Optional[Any]:
        if not self.enable:
            return None
        key = self._key(file_path, extra)
        f = self.cache_dir / f"{key}.pkl.gz"
        if not f.exists():
            return None
        try:
            with gzip.open(f, "rb") as fp:
                return pickle.load(fp)
        except Exception as e:
            logging.info("Cache read failed (%s): %s", f.name, e)
            try:
                f.unlink(missing_ok=True)
            except Exception:
                pass
            return None

    def set(self, file_path: str, extra: str, value: Any) -> None:
        if not self.enable:
            return
        key = self._key(file_path, extra)
        f = self.cache_dir / f"{key}.pkl.gz"
        try:
            with gzip.open(f, "wb") as fp:
                pickle.dump(value, fp, protocol=pickle.HIGHEST_PROTOCOL)
        except Exception as e:
            logging.info("Cache write failed (%s): %s", f.name, e)
        self._trim()

    def clear(self) -> None:
        try:
            for p in self.cache_dir.glob("*.pkl.gz"):
                p.unlink(missing_ok=True)
        except Exception as e:
            logging.warning("Failed to clear cache: %s", e)

    def _trim(self):
        try:
            files = sorted(self.cache_dir.glob("*.pkl.gz"), key=lambda p: p.stat().st_mtime, reverse=True)
            for p in files[self.max_files :]:
                p.unlink(missing_ok=True)
        except Exception:
            pass


# -------------------- Helpers --------------------


_word_re = re.compile(r"[A-Za-z0-9_]+|[\u4e00-\u9fff]+")


def _is_ascii_word(s: str) -> bool:
    return bool(re.fullmatch(r"[A-Za-z0-9_]+", s or ""))


def _eq(a: str, b: str, case_sensitive: bool) -> bool:
    if case_sensitive:
        return (a or "") == (b or "")
    return (a or "").lower() == (b or "").lower()


def _whole_word_accept(keyword: str, detected_word: str, case_sensitive: bool, enabled: bool) -> bool:
    if not enabled:
        return True
    if not _is_ascii_word(keyword):
        return True
    if not detected_word:
        return False
    return _eq(keyword, detected_word, case_sensitive)


def detect_token(line: str, kw: str) -> str:
    if not line or not kw:
        return kw
    idx = line.lower().find(kw.lower())
    if idx < 0:
        return kw
    tokens = []
    for m in _word_re.finditer(line):
        tokens.append((m.start(), m.end(), m.group(0)))
    span = (idx, idx + len(kw))
    for s, e, tok in tokens:
        if s <= span[0] < e or s < span[1] <= e or (span[0] <= s and e <= span[1]):
            return tok
    return kw

def find_plain_occurrences(line: str, kw: str, case_sensitive: bool) -> List[Tuple[int, int, str]]:
    if not line or not kw:
        return []
    hay = line if case_sensitive else line.lower()
    needle = kw if case_sensitive else kw.lower()
    out: List[Tuple[int, int, str]] = []
    start = 0
    while True:
        idx = hay.find(needle, start)
        if idx < 0:
            break
        end = idx + len(kw)
        token = kw
        for m in _word_re.finditer(line):
            s, e, tok = m.start(), m.end(), m.group(0)
            if s <= idx < e or s < end <= e or (idx <= s and e <= end):
                token = tok
                break
        out.append((idx, end, token))
        start = idx + max(1, len(needle))
    return out

def row_values_to_text(values: List[Any]) -> str:
    parts: List[str] = []
    for v in values:
        if v is None:
            continue
        s = str(v).strip()
        if s:
            parts.append(s)
    return " | ".join(parts)

def nonpdf_context(lines: List[str], line_no_1based: int, window: int = 1) -> str:
    if not lines:
        return ""
    i = max(1, min(line_no_1based, len(lines)))
    start = max(1, i - window)
    end = min(len(lines), i + window)
    out = []
    for ln in range(start, end + 1):
        prefix = ">> " if ln == i else "   "
        out.append(prefix + (lines[ln - 1].strip()))
    return "\n".join(out).strip()


def build_line_index(words) -> Tuple[Dict[Tuple[int, int], str], List[Tuple[int, int]], Dict[Tuple[int, int], int]]:
    if not words:
        return {}, [], {}
    tmp: Dict[Tuple[int, int], List[Tuple[int, float, float, str]]] = {}
    line_pos: Dict[Tuple[int, int], Tuple[float, float]] = {}
    for w in words:
        if len(w) < 8:
            continue
        x0, y0, x1, y1, txt, blk, ln, wordno = w[0], w[1], w[2], w[3], w[4], int(w[5]), int(w[6]), int(w[7])
        key = (blk, ln)
        tmp.setdefault(key, []).append((wordno, float(x0), float(y0), str(txt)))
        if key not in line_pos:
            line_pos[key] = (float(y0), float(x0))
        else:
            cy, cx = line_pos[key]
            line_pos[key] = (min(cy, float(y0)), min(cx, float(x0)))
    line_text: Dict[Tuple[int, int], str] = {}
    for key, items in tmp.items():
        items.sort(key=lambda t: (t[0], t[1]))
        line_text[key] = " ".join([t[3] for t in items]).strip()
    ordered_lines = sorted(
        line_text.keys(),
        key=lambda k: (line_pos.get(k, (1e9, 1e9))[0], line_pos.get(k, (1e9, 1e9))[1], k[0], k[1]),
    )
    line_index_map = {k: i for i, k in enumerate(ordered_lines)}
    return line_text, ordered_lines, line_index_map

def build_pdf_word_lookup(words, case_sensitive: bool) -> Dict[str, List[Tuple[str, int, int, Tuple[float, float, float, float]]]]:
    lookup: Dict[str, List[Tuple[str, int, int, Tuple[float, float, float, float]]]] = {}
    if not words:
        return lookup
    for w in words:
        if len(w) < 8:
            continue
        x0, y0, x1, y1, txt, blk, ln = w[0], w[1], w[2], w[3], str(w[4]), int(w[5]), int(w[6])
        key = txt if case_sensitive else txt.lower()
        lookup.setdefault(key, []).append((txt, blk, ln, (float(x0), float(y0), float(x1), float(y1))))
    return lookup

def context_from_line_key(
    line_text: Dict[Tuple[int, int], str],
    ordered_lines: List[Tuple[int, int]],
    line_index_map: Dict[Tuple[int, int], int],
    key: Tuple[int, int],
    window: int = 1,
) -> str:
    if not line_text or not ordered_lines:
        return ""
    idx = line_index_map.get(key)
    if idx is None:
        return ""
    start = max(0, idx - window)
    end = min(len(ordered_lines), idx + window + 1)
    out = []
    for i in range(start, end):
        k = ordered_lines[i]
        prefix = ">> " if k == key else "   "
        out.append(prefix + line_text.get(k, ""))
    return "\n".join(out).strip()

def normalize_line_index_cache(li):
    if not li:
        return {}, [], {}
    if isinstance(li, tuple):
        if len(li) == 3:
            line_text, ordered_lines, line_index_map = li
            if line_index_map is None:
                line_index_map = {k: i for i, k in enumerate(ordered_lines or [])}
            return line_text or {}, ordered_lines or [], line_index_map or {}
        if len(li) == 2:
            line_text, ordered_lines = li
            line_text = line_text or {}
            ordered_lines = ordered_lines or []
            line_index_map = {k: i for i, k in enumerate(ordered_lines)}
            return line_text, ordered_lines, line_index_map
    return {}, [], {}

def best_word_entry_from_rect(words, rect: fitz.Rect) -> Optional[Tuple[str, int, int]]:
    if not words:
        return None
    best = None
    best_score = 0.0
    for w in words:
        if len(w) < 8:
            continue
        x0, y0, x1, y1, txt, blk, ln = w[0], w[1], w[2], w[3], w[4], int(w[5]), int(w[6])
        wr = fitz.Rect(x0, y0, x1, y1)
        inter = rect & wr
        if inter.is_empty:
            continue
        score = max(0.0, inter.get_area())
        if score > best_score:
            best_score = score
            best = (str(txt), blk, ln)
    return best


def seq_ratio(a: str, b: str) -> float:
    # stdlib fuzzy
    from difflib import SequenceMatcher

    if not a or not b:
        return 0.0
    return SequenceMatcher(None, a, b).ratio()


def find_sumatra() -> Optional[str]:
    if not _is_windows():
        return None
    # PATH
    for p in os.environ.get("PATH", "").split(os.pathsep):
        exe = Path(p) / "SumatraPDF.exe"
        if exe.exists():
            return str(exe)
    # common install paths
    candidates = [
        Path(os.environ.get("ProgramFiles", "C:\\Program Files")) / "SumatraPDF" / "SumatraPDF.exe",
        Path(os.environ.get("ProgramFiles(x86)", "C:\\Program Files (x86)")) / "SumatraPDF" / "SumatraPDF.exe",
    ]
    for exe in candidates:
        if exe.exists():
            return str(exe)
    return None


# -------------------- Dialogs --------------------


class FileManagerDialog(QDialog):
    def __init__(self, parent, files: List[str], start_dir: str = ""):
        super().__init__(parent)
        self.setWindowTitle("Files")
        self.setMinimumWidth(780)
        self._files = list(files)
        self._start_dir = start_dir

        layout = QVBoxLayout(self)
        self.list = QListWidget()
        self.list.setSelectionMode(QAbstractItemView.ExtendedSelection)
        layout.addWidget(self.list, 1)

        btn_row = QHBoxLayout()
        b_add = QPushButton("Add…")
        b_remove = QPushButton("Remove selected")
        b_clear = QPushButton("Clear all")
        btn_row.addWidget(b_add)
        btn_row.addWidget(b_remove)
        btn_row.addWidget(b_clear)
        btn_row.addStretch(1)
        layout.addLayout(btn_row)

        self.buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        layout.addWidget(self.buttons)

        b_add.clicked.connect(self._add)
        b_remove.clicked.connect(self._remove)
        b_clear.clicked.connect(self._clear)
        self.buttons.accepted.connect(self.accept)
        self.buttons.rejected.connect(self.reject)

        self._refresh()

    def _refresh(self):
        self.list.clear()
        for p in self._files:
            self.list.addItem(QListWidgetItem(p))

    def _add(self):
        paths, _ = QFileDialog.getOpenFileNames(
            self,
            "Select files",
            self._start_dir or "",
            "All files (*.*);;PDF (*.pdf);;Word (*.docx);;PowerPoint (*.pptx);;Excel (*.xlsx *.xlsm)",
        )
        if not paths:
            return
        for p in paths:
            p = str(Path(p))
            if p not in self._files:
                self._files.append(p)
        self._refresh()

    def _remove(self):
        items = self.list.selectedItems()
        if not items:
            return
        remove = {it.text() for it in items}
        self._files = [f for f in self._files if f not in remove]
        self._refresh()

    def _clear(self):
        self._files = []
        self._refresh()

    def get_files(self) -> List[str]:
        return list(self._files)


class KeywordManagerDialog(QDialog):
    def __init__(self, parent, raw_text: str, cfg: AppConfig):
        super().__init__(parent)
        self.setWindowTitle("Keywords")
        self.setMinimumWidth(720)
        self._cfg = cfg

        layout = QVBoxLayout(self)
        hint = QLabel("One keyword per line is recommended. You can also use comma-separated list.")
        hint.setWordWrap(True)
        layout.addWidget(hint)

        self.box = QTextEdit()
        self.box.setPlainText(raw_text or "")
        self.box.setMinimumHeight(240)
        layout.addWidget(self.box, 1)

        opts = QHBoxLayout()
        self.case_cb = QCheckBox("Case sensitive (English)")
        self.case_cb.setChecked(cfg.case_sensitive)
        self.ww_cb = QCheckBox("Whole word (English)")
        self.ww_cb.setChecked(cfg.whole_word)
        self.mode = QComboBox()
        self.mode.addItems(["plain", "regex", "fuzzy"])
        self.mode.setCurrentText(cfg.search_mode)
        self.fuzzy = QLineEdit(str(cfg.fuzzy_threshold))
        self.fuzzy.setMaximumWidth(80)

        opts.addWidget(self.case_cb)
        opts.addWidget(self.ww_cb)
        opts.addStretch(1)
        opts.addWidget(QLabel("Mode:"))
        opts.addWidget(self.mode)
        opts.addWidget(QLabel("Fuzzy≥"))
        opts.addWidget(self.fuzzy)
        layout.addLayout(opts)

        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        layout.addWidget(buttons)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)

    def get_text(self) -> str:
        return self.box.toPlainText()

    def apply_to_config(self, cfg: AppConfig) -> None:
        cfg.case_sensitive = self.case_cb.isChecked()
        cfg.whole_word = self.ww_cb.isChecked()
        cfg.search_mode = self.mode.currentText().strip() or "plain"
        try:
            v = float(self.fuzzy.text().strip())
            cfg.fuzzy_threshold = max(0.5, min(0.99, v))
        except Exception:
            pass


class SettingsDialog(QDialog):
    def __init__(self, parent, cfg: AppConfig, paths: AppPaths):
        super().__init__(parent)
        self.setWindowTitle("Settings")
        self.setMinimumWidth(560)
        self.cfg = cfg
        self.paths = paths

        layout = QVBoxLayout(self)
        form = QFormLayout()

        self.external = QComboBox()
        self.external.addItems(["auto", "sumatra", "default"])
        self.external.setCurrentText(cfg.external_viewer)

        self.cache_cb = QCheckBox("Enable disk cache")
        self.cache_cb.setChecked(cfg.enable_cache)

        self.cache_info = QLabel(f"Cache dir: {paths.cache_dir}")
        self.cache_info.setTextInteractionFlags(Qt.TextSelectableByMouse)

        form.addRow("External PDF viewer:", self.external)
        form.addRow("", self.cache_cb)
        form.addRow("", self.cache_info)
        layout.addLayout(form)

        btns = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        btns.accepted.connect(self.accept)
        btns.rejected.connect(self.reject)
        layout.addWidget(btns)

    def apply(self):
        self.cfg.external_viewer = self.external.currentText().strip() or "auto"
        self.cfg.enable_cache = self.cache_cb.isChecked()


# -------------------- Preview popup --------------------


class _CtrlWheelZoomFilter(QObject):
    def __init__(self, cb):
        super().__init__()
        self.cb = cb

    def eventFilter(self, obj, event):
        if event.type() == QEvent.Wheel and (event.modifiers() & Qt.ControlModifier):
            delta = event.angleDelta().y()
            if delta > 0:
                self.cb(+1)
            elif delta < 0:
                self.cb(-1)
            return True
        return False


class PreviewPopup(QDialog):
    request_save_clip = Signal()
    request_continue_reading = Signal()
    request_prev_hit = Signal()
    request_next_hit = Signal()

    def __init__(self, parent=None, zoom_default: float = 2.0):
        super().__init__(parent)
        self.setWindowTitle("Preview")
        self.setMinimumSize(980, 760)
        self.setWindowFlag(Qt.Window)

        self.current_hit: Optional[Hit] = None
        self.zoom: float = float(zoom_default)
        self._last_base_size: Optional[Tuple[int, int]] = None

        # caches (in-memory only)
        self._doc_cache: Dict[str, fitz.Document] = {}
        self._page_cache: Dict[Tuple[str, int, int], Image.Image] = {}
        self._doc_lru: List[str] = []
        self._page_lru: List[Tuple[str, int, int]] = []
        self._doc_cache_max = 6
        self._page_cache_max = 24

        self._resize_timer = QTimer(self)
        self._resize_timer.setSingleShot(True)
        self._resize_timer.timeout.connect(self._rerender)

        root = QWidget()
        self.setLayout(QVBoxLayout())
        self.layout().setContentsMargins(10, 10, 10, 10)
        self.layout().addWidget(root)

        outer = QVBoxLayout(root)
        outer.setContentsMargins(0, 0, 0, 0)
        outer.setSpacing(10)

        # top controls
        row = QHBoxLayout()
        self.btn_continue = QPushButton("Continue reading")
        self.btn_continue.clicked.connect(lambda: self.request_continue_reading.emit())
        self.btn_prev = QPushButton("◀ Prev")
        self.btn_prev.clicked.connect(lambda: self.request_prev_hit.emit())
        self.btn_next = QPushButton("Next ▶")
        self.btn_next.clicked.connect(lambda: self.request_next_hit.emit())

        row.addWidget(self.btn_continue)
        row.addStretch(1)
        row.addWidget(self.btn_prev)
        row.addWidget(self.btn_next)
        outer.addLayout(row)

        # zoom controls
        zrow = QHBoxLayout()
        self.btn_fit_w = QPushButton("Fit width")
        self.btn_fit_w.clicked.connect(self.fit_width)
        self.btn_fit_p = QPushButton("Fit page")
        self.btn_fit_p.clicked.connect(self.fit_page)
        self.btn_zout = QPushButton("-")
        self.btn_zout.setShortcut(QKeySequence("Ctrl+-"))
        self.btn_zout.clicked.connect(self.zoom_out)
        self.btn_zreset = QPushButton("100%")
        self.btn_zreset.setShortcut(QKeySequence("Ctrl+0"))
        self.btn_zreset.clicked.connect(self.zoom_reset)
        self.btn_zin = QPushButton("+")
        self.btn_zin.setShortcut(QKeySequence("Ctrl+="))
        self.btn_zin.clicked.connect(self.zoom_in)

        zrow.addStretch(1)
        zrow.addWidget(self.btn_fit_w)
        zrow.addWidget(self.btn_fit_p)
        zrow.addWidget(self.btn_zout)
        zrow.addWidget(self.btn_zreset)
        zrow.addWidget(self.btn_zin)
        outer.addLayout(zrow)

        # preview group
        preview_group = QGroupBox("PDF page")
        pg = QVBoxLayout(preview_group)
        self.preview_label = QLabel("Select a PDF hit to preview.")
        self.preview_label.setAlignment(Qt.AlignLeft | Qt.AlignTop)
        self.preview_label.setStyleSheet("background:#111;color:#ddd;padding:10px;")

        self.preview_scroll = QScrollArea()
        self.preview_scroll.setWidgetResizable(False)
        self.preview_scroll.setStyleSheet("QScrollArea { border: 1px solid rgba(0,0,0,0.15); }")
        self.preview_scroll.setWidget(self.preview_label)
        pg.addWidget(self.preview_scroll, 1)
        outer.addWidget(preview_group, 1)

        # bottom text boxes
        ctx = QGroupBox("Context")
        cl = QVBoxLayout(ctx)
        self.context_box = QTextEdit()
        self.context_box.setReadOnly(True)
        self.context_box.setMinimumHeight(120)
        cl.addWidget(self.context_box)
        outer.addWidget(ctx)

        txt = QGroupBox("Page text (select to clip)")
        tl = QVBoxLayout(txt)
        top = QHBoxLayout()
        top.addStretch(1)
        self.btn_save_clip = QPushButton("Save selection as clip")
        self.btn_save_clip.clicked.connect(lambda: self.request_save_clip.emit())
        top.addWidget(self.btn_save_clip)
        tl.addLayout(top)
        self.page_text_box = QTextEdit()
        self.page_text_box.setReadOnly(True)
        self.page_text_box.setMinimumHeight(160)
        tl.addWidget(self.page_text_box)
        outer.addWidget(txt)

        # Ctrl+Wheel zoom
        self._zoom_filter = _CtrlWheelZoomFilter(self._on_ctrl_wheel)
        self.preview_scroll.viewport().installEventFilter(self._zoom_filter)
        self.preview_label.installEventFilter(self._zoom_filter)

    def _on_ctrl_wheel(self, direction: int):
        if direction > 0:
            self.zoom_in()
        else:
            self.zoom_out()

    def closeEvent(self, event):
        try:
            for d in self._doc_cache.values():
                try:
                    d.close()
                except Exception:
                    pass
        except Exception:
            pass
        self._doc_cache.clear()
        self._page_cache.clear()
        super().closeEvent(event)

    def resizeEvent(self, event):
        super().resizeEvent(event)
        self._resize_timer.start(120)

    # ----- cache helpers -----
    def _touch_doc(self, key: str):
        if key in self._doc_lru:
            self._doc_lru.remove(key)
        self._doc_lru.append(key)
        while len(self._doc_lru) > self._doc_cache_max:
            old = self._doc_lru.pop(0)
            d = self._doc_cache.pop(old, None)
            if d is not None:
                try:
                    d.close()
                except Exception:
                    pass

    def _touch_page(self, key: Tuple[str, int, int]):
        if key in self._page_lru:
            self._page_lru.remove(key)
        self._page_lru.append(key)
        while len(self._page_lru) > self._page_cache_max:
            old = self._page_lru.pop(0)
            self._page_cache.pop(old, None)

    def _get_doc_cached(self, pdf_path: str) -> fitz.Document:
        key = str(Path(pdf_path).resolve())
        d = self._doc_cache.get(key)
        if d is None:
            d = fitz.open(key)
            self._doc_cache[key] = d
        self._touch_doc(key)
        return d

    def _get_page_image_cached(self, pdf_path: str, page_1based: int, zoom: float) -> Image.Image:
        key = (str(Path(pdf_path).resolve()), int(page_1based), int(round(zoom * 100)))
        img = self._page_cache.get(key)
        if img is not None:
            self._touch_page(key)
            return img
        doc = self._get_doc_cached(pdf_path)
        page = doc.load_page(page_1based - 1)
        mat = fitz.Matrix(zoom, zoom)
        pix = page.get_pixmap(matrix=mat, alpha=False)
        img = Image.frombytes("RGB", (pix.width, pix.height), pix.samples)
        self._page_cache[key] = img
        self._touch_page(key)
        return img

    # ----- zoom -----
    def set_zoom(self, zoom: float):
        self.zoom = max(0.5, min(8.0, float(zoom)))
        self._rerender()

    def zoom_in(self):
        self.set_zoom(self.zoom * 1.25)

    def zoom_out(self):
        self.set_zoom(self.zoom / 1.25)

    def zoom_reset(self):
        self.set_zoom(1.0)

    def fit_width(self):
        hit = self.current_hit
        if not hit or hit.file_type != "PDF" or not hit.page:
            return
        try:
            # render once at current zoom; base size is tracked
            doc = self._get_doc_cached(hit.file_path)
            page = doc.load_page(hit.page - 1)
            # compute zoom so that page width fits viewport
            w = page.rect.width
            viewport = self.preview_scroll.viewport().width() - 20
            if w > 0:
                self.set_zoom(max(0.5, min(8.0, viewport / w)))
        except Exception:
            pass

    def fit_page(self):
        hit = self.current_hit
        if not hit or hit.file_type != "PDF" or not hit.page:
            return
        try:
            doc = self._get_doc_cached(hit.file_path)
            page = doc.load_page(hit.page - 1)
            w = page.rect.width
            h = page.rect.height
            vw = self.preview_scroll.viewport().width() - 20
            vh = self.preview_scroll.viewport().height() - 20
            if w > 0 and h > 0:
                self.set_zoom(max(0.5, min(8.0, min(vw / w, vh / h))))
        except Exception:
            pass

    # ----- main -----
    def show_hit(self, hit: Hit):
        self.current_hit = hit
        self.context_box.setPlainText(hit.context or hit.snippet or "")
        if hit.file_type == "PDF" and hit.page is not None:
            try:
                doc = self._get_doc_cached(hit.file_path)
                page = doc.load_page(hit.page - 1)
                txt = page.get_text("text") or ""
                self.page_text_box.setPlainText(txt.strip())
            except Exception:
                self.page_text_box.setPlainText(hit.context or hit.snippet or "")
        else:
            self.page_text_box.setPlainText(hit.context or hit.snippet or "")
        self._rerender()

    def _rerender(self):
        hit = self.current_hit
        if not hit:
            return
        if hit.file_type != "PDF" or hit.page is None or hit.rect is None:
            self.preview_label.setPixmap(QPixmap())
            self.preview_label.setText("(No PDF image preview for this hit.)")
            self.preview_label.adjustSize()
            return
        try:
            base = self._get_page_image_cached(hit.file_path, hit.page, self.zoom)
            img = base.copy()
            draw = ImageDraw.Draw(img)
            x0, y0, x1, y1 = hit.rect
            rr = fitz.Rect(x0 * self.zoom, y0 * self.zoom, x1 * self.zoom, y1 * self.zoom)
            draw.rectangle([rr.x0, rr.y0, rr.x1, rr.y1], outline="red", width=max(2, int(3 * self.zoom)))
            qimg = ImageQt.ImageQt(img)
            pm = QPixmap.fromImage(qimg)
            self.preview_label.setText("")
            self.preview_label.setPixmap(pm)
            self.preview_label.setFixedSize(pm.size())
        except Exception as e:
            self.preview_label.setPixmap(QPixmap())
            self.preview_label.setText(f"Failed to render preview:\n{e}")
            self.preview_label.adjustSize()

    def get_selected_text(self) -> str:
        for w in (self.page_text_box, self.context_box):
            try:
                cur = w.textCursor()
                if cur is not None and cur.hasSelection():
                    return (cur.selectedText() or "").replace("\u2029", "\n").strip()
            except Exception:
                pass
        return ""


# -------------------- Clips dialog --------------------


class ClipsDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Clips")
        self.setMinimumSize(860, 560)
        self.setWindowFlag(Qt.Window)
        self.clips: List[Clip] = []

        root = QWidget()
        self.setLayout(QVBoxLayout())
        self.layout().setContentsMargins(10, 10, 10, 10)
        self.layout().addWidget(root)

        outer = QVBoxLayout(root)
        self.table = QTableWidget(0, 5)
        self.table.setHorizontalHeaderLabels(["Created", "File", "Page/Line", "Keyword", "Preview"])
        self.table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(4, QHeaderView.Stretch)
        self.table.verticalHeader().setVisible(False)
        self.table.setAlternatingRowColors(True)
        self.table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.table.setSelectionBehavior(QTableWidget.SelectRows)
        self.table.setSelectionMode(QAbstractItemView.SingleSelection)
        outer.addWidget(self.table, 1)

        self.detail = QTextEdit()
        self.detail.setReadOnly(True)
        self.detail.setMinimumHeight(150)
        outer.addWidget(self.detail)

        bb = QDialogButtonBox(QDialogButtonBox.Close)
        bb.rejected.connect(self.close)
        bb.accepted.connect(self.close)
        outer.addWidget(bb)

        self.table.cellClicked.connect(self._on_row)

    def refresh(self, clips: List[Clip]):
        self.clips = list(clips)
        self.table.setRowCount(len(self.clips))
        for i, c in enumerate(self.clips):
            where = ""
            if c.file_type == "PDF" and c.page is not None:
                where = f"p.{c.page}"
            elif c.row_no is not None:
                where = f"row.{c.row_no}"    
            elif c.line_no is not None:
                where = f"ln.{c.line_no}"
            preview = (c.selected_text or "").replace("\n", " ").strip()
            if len(preview) > 160:
                preview = preview[:160] + "…"
            self.table.setItem(i, 0, QTableWidgetItem(c.created_at))
            self.table.setItem(i, 1, QTableWidgetItem(Path(c.file_path).name))
            self.table.setItem(i, 2, QTableWidgetItem(where))
            self.table.setItem(i, 3, QTableWidgetItem(c.keyword))
            self.table.setItem(i, 4, QTableWidgetItem(preview))
        if not self.clips:
            self.detail.setPlainText("(No clips yet. Select text in Preview and click 'Save selection as clip'.)")

    def select_last(self):
        if not self.clips:
            return
        last = len(self.clips) - 1
        self.table.selectRow(last)
        self.table.scrollToItem(self.table.item(last, 0), QAbstractItemView.PositionAtCenter)
        self._on_row(last, 0)

    def _on_row(self, row: int, col: int):
        if row < 0 or row >= len(self.clips):
            return
        c = self.clips[row]
        where = ""
        if c.file_type == "PDF" and c.page is not None:
            where = f"Page {c.page}"
        elif c.row_no is not None:
            where = f"Row {c.row_no}"
        elif c.line_no is not None:
            where = f"Line {c.line_no}"
        self.detail.setPlainText(
            f"{Path(c.file_path).name} — {where}\n"
            f"Keyword: {c.keyword} (detected: {c.detected_word})\n"
            f"Saved: {c.created_at}\n\n"
            f"{c.selected_text}"
        )


# -------------------- Worker --------------------


class SearchWorker(QThread):
    progress_text = Signal(str)
    progress_range = Signal(int, int)  # min, max
    progress_value = Signal(int)
    finished_ok = Signal(list)  # List[Hit]
    finished_err = Signal(str)
    cancelled = Signal()

    def __init__(
        self,
        files: List[str],
        keywords: List[str],
        cfg: AppConfig,
        cache: DiskCache,
    ):
        super().__init__()
        self.files = files
        self.keywords = keywords
        self.cfg = cfg
        self.cache = cache
        self._stop = False

    def stop(self):
        self._stop = True

    def _check_stop(self) -> bool:
        if self._stop:
            self.progress_text.emit("Cancelled.")
            self.cancelled.emit()
            return True
        return False
    def _append_sheet_hits(self, hits: List[Hit], fpath: str, file_type: str, row_no: int, row_text: str, keyword_entries: List[Dict[str, Any]]):
        if not row_text:
            return

        lines_for_context = [row_text]
        for entry in keyword_entries:
            kw = entry["kw"]
            kw_cmp = entry["kw_cmp"]

            if self.cfg.search_mode == "regex":
                try:
                    pat = re.compile(kw, 0 if self.cfg.case_sensitive else re.IGNORECASE)
                except re.error:
                    continue
                matches = list(pat.finditer(row_text))
                if not matches:
                    continue
                ctx = nonpdf_context(lines_for_context, 1, window=0)
                for m in matches:
                    detected = (m.group(0) or kw).strip()
                    if not _whole_word_accept(kw, detected, self.cfg.case_sensitive, self.cfg.whole_word):
                        continue
                    hits.append(
                        Hit(
                            file_path=fpath,
                            file_type=file_type,
                            keyword=kw,
                            detected_word=detected,
                            page=None,
                            line_no=None,
                            row_no=row_no,
                            rect=None,
                            snippet=row_text,
                            context=ctx,
                        )
                    )
                continue

            if self.cfg.search_mode == "fuzzy":
                line_tokens = [t.group(0) for t in _word_re.finditer(row_text)]
                line_tokens_cmp = [tok if self.cfg.case_sensitive else tok.lower() for tok in line_tokens]

                best_tok = ""
                best_sc = 0.0
                kw_len = len(kw_cmp)

                for tok, tok_cmp in zip(line_tokens, line_tokens_cmp):
                    if abs(len(tok_cmp) - kw_len) > max(2, int(kw_len * 0.4)):
                        continue
                    sc = seq_ratio(kw_cmp, tok_cmp)
                    if sc > best_sc:
                        best_sc, best_tok = sc, tok

                if best_sc < float(self.cfg.fuzzy_threshold):
                    continue
                detected = best_tok or kw
                if not _whole_word_accept(kw, detected, self.cfg.case_sensitive, self.cfg.whole_word):
                    continue

                ctx = nonpdf_context(lines_for_context, 1, window=0)
                hits.append(
                    Hit(
                        file_path=fpath,
                        file_type=file_type,
                        keyword=kw,
                        detected_word=detected,
                        page=None,
                        line_no=None,
                        row_no=row_no,
                        rect=None,
                        snippet=row_text,
                        context=ctx,
                    )
                )
                continue

            occurrences = find_plain_occurrences(row_text, kw, self.cfg.case_sensitive)
            if not occurrences:
                continue

            ctx = nonpdf_context(lines_for_context, 1, window=0)
            for _start, _end, detected in occurrences:
                if not _whole_word_accept(kw, detected, self.cfg.case_sensitive, self.cfg.whole_word):
                    continue
                hits.append(
                    Hit(
                        file_path=fpath,
                        file_type=file_type,
                        keyword=kw,
                        detected_word=detected or kw,
                        page=None,
                        line_no=None,
                        row_no=row_no,
                        rect=None,
                        snippet=row_text,
                        context=ctx,
                    )
                )

    def _search_csv_file(self, fpath: str, hits: List[Hit], keyword_entries: List[Dict[str, Any]]):
        encodings = ["utf-8-sig", "utf-8", "gb18030", "latin-1"]
        last_err = None
        for enc in encodings:
            try:
                with open(fpath, "r", encoding=enc, newline="") as fp:
                    reader = csv.reader(fp)
                    for row_no, row in enumerate(reader, 1):
                        row_text = row_values_to_text(row)
                        if row_text:
                            self._append_sheet_hits(hits, fpath, "CSV", row_no, row_text, keyword_entries)
                return
            except Exception as e:
                last_err = e
        if last_err is not None:
            raise last_err

    def _search_excel_file(self, fpath: str, hits: List[Hit], keyword_entries: List[Dict[str, Any]]):
        wb = load_workbook(fpath, read_only=True, data_only=True)
        try:
            for ws in wb.worksheets:
                for row_no, row in enumerate(ws.iter_rows(values_only=True), 1):
                    row_text = row_values_to_text(list(row))
                    if row_text:
                        self._append_sheet_hits(hits, fpath, "XLSX", row_no, row_text, keyword_entries)
        finally:
            wb.close()

    def run(self):
        try:
            if not self.files:
                self.finished_err.emit("No files selected.")
                return
            if not self.keywords:
                self.finished_err.emit("No keywords provided.")
                return
            keyword_entries = []
            for kw in self.keywords:
                if not kw:
                    continue
                keyword_entries.append(
                    {
                        "kw": kw,
                        "kw_cmp": kw if self.cfg.case_sensitive else kw.lower(),
                        "is_ascii_word": _is_ascii_word(kw),
                    }
                )
            if not keyword_entries:
                self.finished_err.emit("No valid keywords provided.")
                return

            pdfs = [f for f in self.files if Path(f).suffix.lower() == ".pdf"]
            spreadsheets = [
                f for f in self.files
                if Path(f).suffix.lower() in {".csv", ".xlsx", ".xlsm", ".xltx", ".xltm"}
            ]
            others = [
                f for f in self.files
                if Path(f).suffix.lower() != ".pdf"
                and Path(f).suffix.lower() not in {".csv", ".xlsx", ".xlsm", ".xltx", ".xltm"}
            ]

            hits: List[Hit] = []

            # progress by file
            total_files = len(pdfs) + len(spreadsheets) + len(others)
            self.progress_range.emit(0, max(1, total_files))
            done_files = 0

            # --- PDF search ---
            for fpath in pdfs:
                if self._check_stop():
                    return
                done_files += 1
                self.progress_value.emit(done_files)
                self.progress_text.emit(f"PDF: {Path(fpath).name} ({done_files}/{total_files})")

                try:
                    doc = fitz.open(fpath)
                except Exception as e:
                    self.progress_text.emit(f"Failed to open PDF: {Path(fpath).name} — {e}")
                    continue

                page_count = len(doc)
                for pno in range(page_count):
                    if self._check_stop():
                        try:
                            doc.close()
                        except Exception:
                            pass
                        return

                    page = doc.load_page(pno)
                    page_text = None
                    words = None
                    line_text = None
                    ordered_lines = None
                    line_index_map = None
                    pdf_lookup = None

                    # do the search first; build heavier page structures only when needed
                    for entry in keyword_entries:
                        kw = entry["kw"]

                        if self.cfg.search_mode == "regex":
                            # best-effort regex on extracted page text (no true bbox guarantee)
                            try:
                                pat = re.compile(kw, 0 if self.cfg.case_sensitive else re.IGNORECASE)
                            except re.error:
                                continue
                            page_text = page.get_text("text") or ""
                            for m in pat.finditer(page_text):
                                detected = (m.group(0) or kw).strip()
                                # bbox: try search_for detected literal
                                rects = page.search_for(detected, flags=fitz.TEXTFLAGS_SEARCH) if detected else []
                                r = rects[0] if rects else None
                                rect_tuple = (r.x0, r.y0, r.x1, r.y1) if r else None
                                # context/snippet: best-effort via cached line index
                                li = self.cache.get(fpath, extra=f"pdf_idx_{pno}")
                                if li is None:
                                    words = page.get_text("words") or []
                                    line_text, ordered_lines, line_index_map = build_line_index(words)
                                    li = (line_text, ordered_lines, line_index_map)
                                    self.cache.set(fpath, extra=f"pdf_idx_{pno}", value=li)
                                line_text, ordered_lines, line_index_map = normalize_line_index_cache(li)
                                snippet = detected
                                ctx = detected
                                hits.append(
                                    Hit(
                                        file_path=fpath,
                                        file_type="PDF",
                                        keyword=kw,
                                        detected_word=detected or kw,
                                        page=pno + 1,
                                        line_no=None,
                                        row_no=None,
                                        rect=rect_tuple,
                                        snippet=snippet,
                                        context=ctx,
                                    )
                                )
                        else:
                            use_word_lookup = self.cfg.whole_word and entry["is_ascii_word"]

                            if use_word_lookup:
                                if words is None:
                                    words = self.cache.get(fpath, extra=f"pdf_words_{pno}")
                                li = None if line_text is not None else self.cache.get(fpath, extra=f"pdf_idx_{pno}")
                                lookup_extra = f"pdf_lookup_cs_{int(self.cfg.case_sensitive)}_{pno}"
                                if pdf_lookup is None:
                                    pdf_lookup = self.cache.get(fpath, extra=lookup_extra)

                                if words is None or li is None or pdf_lookup is None:
                                    if words is None:
                                        words = page.get_text("words") or []
                                        self.cache.set(fpath, extra=f"pdf_words_{pno}", value=words)
                                    line_text, ordered_lines, line_index_map = build_line_index(words)
                                    li = (line_text, ordered_lines, line_index_map)
                                    self.cache.set(fpath, extra=f"pdf_idx_{pno}", value=li)
                                    pdf_lookup = build_pdf_word_lookup(words, self.cfg.case_sensitive)
                                    self.cache.set(fpath, extra=lookup_extra, value=pdf_lookup)
                                else:
                                    line_text, ordered_lines, line_index_map = normalize_line_index_cache(li)

                                lookup_key = entry["kw_cmp"]
                                matches = pdf_lookup.get(lookup_key, [])
                                if not matches:
                                    continue

                                for detected, blk, ln, rect_tuple in matches:
                                    snippet = line_text.get((blk, ln), "")
                                    ctx = context_from_line_key(line_text, ordered_lines, line_index_map, (blk, ln), window=1)
                                    hits.append(
                                        Hit(
                                            file_path=fpath,
                                            file_type="PDF",
                                            keyword=kw,
                                            detected_word=detected or kw,
                                            page=pno + 1,
                                            line_no=None,
                                            row_no=None,
                                            rect=rect_tuple,
                                            snippet=snippet,
                                            context=ctx or snippet,
                                        )
                                    )
                            else:
                                rects = page.search_for(kw, flags=fitz.TEXTFLAGS_SEARCH)
                                if not rects:
                                    continue

                                if words is None:
                                    words = self.cache.get(fpath, extra=f"pdf_words_{pno}")
                                li = None if line_text is not None else self.cache.get(fpath, extra=f"pdf_idx_{pno}")
                                if words is None or li is None:
                                    if words is None:
                                        words = page.get_text("words") or []
                                        self.cache.set(fpath, extra=f"pdf_words_{pno}", value=words)
                                    line_text, ordered_lines, line_index_map = build_line_index(words)
                                    li = (line_text, ordered_lines, line_index_map)
                                    self.cache.set(fpath, extra=f"pdf_idx_{pno}", value=li)
                                else:
                                    line_text, ordered_lines, line_index_map = normalize_line_index_cache(li)

                                for r in rects:
                                    best = best_word_entry_from_rect(words, r)
                                    if best is None:
                                        detected, blk, ln = kw, -1, -1
                                        snippet = ""
                                        ctx = ""
                                    else:
                                        detected, blk, ln = best
                                        snippet = line_text.get((blk, ln), "")
                                        ctx = context_from_line_key(line_text, ordered_lines, line_index_map, (blk, ln), window=1)

                                    if not _whole_word_accept(kw, detected, self.cfg.case_sensitive, self.cfg.whole_word):
                                        continue

                                    hits.append(
                                        Hit(
                                            file_path=fpath,
                                            file_type="PDF",
                                            keyword=kw,
                                            detected_word=detected or kw,
                                            page=pno + 1,
                                            row_no=None,
                                            line_no=None,
                                            rect=(r.x0, r.y0, r.x1, r.y1),
                                            snippet=snippet,
                                            context=ctx or snippet,
                                        )
                                    )

                try:
                    doc.close()
                except Exception:
                    pass
            # --- Spreadsheet / CSV search ---
            if spreadsheets:
                for fpath in spreadsheets:
                    if self._check_stop():
                        return
                    done_files += 1
                    self.progress_value.emit(min(done_files, total_files))
                    self.progress_text.emit(f"Sheet: {Path(fpath).name} ({done_files}/{total_files})")

                    try:
                        suffix = Path(fpath).suffix.lower()
                        if suffix == ".csv":
                            self._search_csv_file(fpath, hits, keyword_entries)
                        else:
                            self._search_excel_file(fpath, hits, keyword_entries)
                    except Exception as e:
                        self.progress_text.emit(f"Failed to read sheet: {Path(fpath).name} — {e}")
                        continue            
            # --- Non-PDF search ---
            if others:
                if not KREUZBERG_AVAILABLE:
                    self.finished_err.emit(
                        "kreuzberg is not installed, so non-PDF search (docx/xlsx/pptx/…) is unavailable.\n\n"
                        "Fix: pip install kreuzberg"
                    )
                    return
                if self._check_stop():
                    return
                # extract (cached)
                self.progress_text.emit("Extracting non-PDF files with kreuzberg…")
                cfg = ExtractionConfig()
                to_extract = []
                extracted_map: Dict[str, str] = {}
                for fpath in others:
                    cached = self.cache.get(fpath, extra="nonpdf_text")
                    if cached is None:
                        to_extract.append(fpath)
                    else:
                        extracted_map[fpath] = cached

                if to_extract:
                    results = batch_extract_files_sync(to_extract, config=cfg)
                    for i, result in enumerate(results):
                        fpath = to_extract[i]
                        content = getattr(result, "content", "") or ""
                        extracted_map[fpath] = content
                        self.cache.set(fpath, extra="nonpdf_text", value=content)

                for fpath in others:
                    if self._check_stop():
                        return
                    done_files += 1
                    self.progress_value.emit(min(done_files, total_files))
                    self.progress_text.emit(f"Text: {Path(fpath).name} ({done_files}/{total_files})")

                    content = extracted_map.get(fpath, "") or ""
                    lines = content.splitlines()

                    for ln, raw_line in enumerate(lines, 1):
                        if self._check_stop():
                            return
                        line = raw_line.strip()
                        if not line:
                            continue

                        line_cmp = line if self.cfg.case_sensitive else line.lower()
                        line_tokens = None
                        line_tokens_cmp = None

                        for entry in keyword_entries:
                            kw = entry["kw"]
                            kw_cmp = entry["kw_cmp"]
                            # match mode
                            if self.cfg.search_mode == "regex":
                                try:
                                    pat = re.compile(kw, 0 if self.cfg.case_sensitive else re.IGNORECASE)
                                except re.error:
                                    continue
                                matches = list(pat.finditer(line))
                                if not matches:
                                    continue
                                ctx = nonpdf_context(lines, ln, window=1)
                                for m in matches:
                                    detected = (m.group(0) or kw).strip()
                                    if not _whole_word_accept(kw, detected, self.cfg.case_sensitive, self.cfg.whole_word):
                                        continue
                                    hits.append(
                                        Hit(
                                            file_path=fpath,
                                            file_type="DOC/PPT/XLS/…",
                                            keyword=kw,
                                            detected_word=detected,
                                            page=None,
                                            line_no=ln,
                                            row_no=None,
                                            rect=None,
                                            snippet=line,
                                            context=ctx,
                                        )
                                    )
                                continue
                            elif self.cfg.search_mode == "fuzzy":
                                # tokenize once per line
                                if line_tokens is None:
                                    line_tokens = [t.group(0) for t in _word_re.finditer(line)]
                                    line_tokens_cmp = [tok if self.cfg.case_sensitive else tok.lower() for tok in line_tokens]

                                best_tok = ""
                                best_sc = 0.0
                                kw_len = len(kw_cmp)

                                for tok, tok_cmp in zip(line_tokens, line_tokens_cmp or []):
                                    # cheap prefilter before SequenceMatcher
                                    if abs(len(tok_cmp) - kw_len) > max(2, int(kw_len * 0.4)):
                                        continue
                                    sc = seq_ratio(kw_cmp, tok_cmp)
                                    if sc > best_sc:
                                        best_sc, best_tok = sc, tok

                                if best_sc < float(self.cfg.fuzzy_threshold):
                                    continue
                                detected = best_tok or kw

                                if not _whole_word_accept(kw, detected, self.cfg.case_sensitive, self.cfg.whole_word):
                                    continue

                                ctx = nonpdf_context(lines, ln, window=1)
                                hits.append(
                                    Hit(
                                        file_path=fpath,
                                        file_type="DOC/PPT/XLS/…",
                                        keyword=kw,
                                        detected_word=detected,
                                        page=None,
                                        line_no=ln,
                                        row_no=None,
                                        rect=None,
                                        snippet=line,
                                        context=ctx,
                                    )
                                )
                            else:
                                occurrences = find_plain_occurrences(line, kw, self.cfg.case_sensitive)
                                if not occurrences:
                                    continue

                                ctx = nonpdf_context(lines, ln, window=1)
                                for _start, _end, detected in occurrences:
                                    if not _whole_word_accept(kw, detected, self.cfg.case_sensitive, self.cfg.whole_word):
                                        continue
                                    hits.append(
                                        Hit(
                                            file_path=fpath,
                                            file_type="DOC/PPT/XLS/…",
                                            keyword=kw,
                                            detected_word=detected or kw,
                                            page=None,
                                            line_no=ln,
                                            row_no=None,
                                            rect=None,
                                            snippet=line,
                                            context=ctx,
                                        )
                                    )
                                continue

            self.progress_text.emit(f"Done. {len(hits)} hit(s).")
            self.finished_ok.emit(hits)
        except Exception as e:
            logging.exception("Worker crash")
            self.finished_err.emit(str(e))


# -------------------- Main window --------------------


class KeywordFinderWindow(QMainWindow):
    def __init__(self, paths: AppPaths, cfg: AppConfig):
        super().__init__()
        self.paths = paths
        self.cfg = cfg
        self.cache = DiskCache(paths.cache_dir, enable=cfg.enable_cache, max_files=cfg.cache_max_files)

        self.setWindowTitle("Keyword Finder")
        self.setMinimumSize(1200, 720)

        self.files: List[str] = []
        self.keywords_text: str = ""
        self.hits: List[Hit] = []
        self._view_rows: List[int] = []  # mapping from table row -> hit index (Hit view) or representative hit
        self.clips: List[Clip] = []
        self.worker: Optional[SearchWorker] = None

        self.preview = PreviewPopup(self, zoom_default=cfg.zoom_default)
        self.preview.request_save_clip.connect(self.save_current_selection_as_clip)
        self.preview.request_continue_reading.connect(self.open_selected_in_viewer)
        self.preview.request_prev_hit.connect(self.select_prev)
        self.preview.request_next_hit.connect(self.select_next)
        self.clips_dialog = ClipsDialog(self)

        self._build_ui()
        self._build_menus()
        self._refresh_status()

    # ----- UI -----
    def _build_ui(self):
        root = QWidget()
        self.setCentralWidget(root)
        layout = QVBoxLayout(root)
        layout.setContentsMargins(10, 10, 10, 10)
        layout.setSpacing(10)

        # top row
        top = QHBoxLayout()
        self.btn_files = QPushButton("Files…")
        self.btn_files.clicked.connect(self.manage_files)
        self.btn_kw = QPushButton("Keywords…")
        self.btn_kw.clicked.connect(self.manage_keywords)
        self.btn_run = QPushButton("Run search")
        self.btn_run.clicked.connect(self.run_search)
        self.btn_cancel = QPushButton("Cancel")
        self.btn_cancel.clicked.connect(self.cancel_search)
        self.btn_cancel.setEnabled(False)
        self.btn_preview = QPushButton("Preview")
        self.btn_preview.clicked.connect(self.toggle_preview)

        top.addWidget(self.btn_files)
        top.addWidget(self.btn_kw)
        top.addWidget(self.btn_run)
        top.addWidget(self.btn_cancel)
        top.addStretch(1)
        top.addWidget(self.btn_preview)
        layout.addLayout(top)

        # results
        g = QGroupBox("Results")
        gl = QVBoxLayout(g)
        self.table = QTableWidget(0, 9)
        self.table.setHorizontalHeaderLabels(["View", "Type", "File", "Keyword", "Detected", "Page", "Line", "Row", "Snippet"])
        self.table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(4, QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(5, QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(6, QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(7, QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(8, QHeaderView.Stretch)
        self.table.verticalHeader().setVisible(False)
        self.table.setAlternatingRowColors(True)
        self.table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.table.setSelectionBehavior(QTableWidget.SelectRows)
        self.table.setSelectionMode(QAbstractItemView.SingleSelection)
        self.table.cellClicked.connect(self._on_row_clicked)
        self.table.cellDoubleClicked.connect(lambda r, c: self.open_selected_in_viewer())
        gl.addWidget(self.table, 1)
        layout.addWidget(g, 1)

        # status + progress
        status_row = QHBoxLayout()
        self.status = QLabel("Ready.")
        self.status.setWordWrap(True)
        self.pbar = QProgressBar()
        self.pbar.setFixedWidth(260)
        self.pbar.setRange(0, 1)
        self.pbar.setValue(0)
        status_row.addWidget(self.status, 1)
        status_row.addWidget(self.pbar)
        layout.addLayout(status_row)

    def _build_menus(self):
        mb = self.menuBar()
        mb.clear()

        m_files = mb.addMenu("Files")
        a_manage_files = QAction("Manage files…", self)
        a_manage_files.setShortcut(QKeySequence("Ctrl+O"))
        a_manage_files.triggered.connect(self.manage_files)
        a_clear_files = QAction("Clear files", self)
        a_clear_files.triggered.connect(self.clear_files)
        m_files.addAction(a_manage_files)
        m_files.addAction(a_clear_files)

        m_kw = mb.addMenu("Keywords")
        a_manage_kw = QAction("Manage keywords…", self)
        a_manage_kw.setShortcut(QKeySequence("Ctrl+K"))
        a_manage_kw.triggered.connect(self.manage_keywords)
        a_import_kw = QAction("Import keywords…", self)
        a_import_kw.triggered.connect(self.import_keywords)
        a_export_kw = QAction("Export keywords…", self)
        a_export_kw.triggered.connect(self.export_keywords)
        a_export_prompt = QAction("Export AI prompt template…", self)
        a_export_prompt.triggered.connect(self.export_prompt_template)
        m_kw.addAction(a_manage_kw)
        m_kw.addSeparator()
        m_kw.addAction(a_import_kw)
        m_kw.addAction(a_export_kw)
        m_kw.addAction(a_export_prompt)

        m_search = mb.addMenu("Search")
        a_run = QAction("Run", self)
        a_run.setShortcut(QKeySequence("Ctrl+R"))
        a_run.triggered.connect(self.run_search)
        a_cancel = QAction("Cancel", self)
        a_cancel.setShortcut(QKeySequence("Esc"))
        a_cancel.triggered.connect(self.cancel_search)
        m_search.addAction(a_run)
        m_search.addAction(a_cancel)

        m_view = mb.addMenu("View")
        self.a_view_hits = QAction("Hit view", self, checkable=True)
        self.a_view_pages = QAction("Page view (grouped)", self, checkable=True)
        self.a_view_hits.setChecked(self.cfg.results_view == "hits")
        self.a_view_pages.setChecked(self.cfg.results_view == "pages")
        self.a_view_hits.triggered.connect(lambda: self.set_results_view("hits"))
        self.a_view_pages.triggered.connect(lambda: self.set_results_view("pages"))
        m_view.addAction(self.a_view_hits)
        m_view.addAction(self.a_view_pages)

        m_preview = mb.addMenu("Preview")
        a_show = QAction("Show/Hide preview", self)
        a_show.setShortcut(QKeySequence("Ctrl+P"))
        a_show.triggered.connect(self.toggle_preview)
        a_open = QAction("Continue reading (open PDF)", self)
        a_open.setShortcut(QKeySequence(Qt.Key_Return))
        a_open.triggered.connect(self.open_selected_in_viewer)
        m_preview.addAction(a_show)
        m_preview.addAction(a_open)

        m_export = mb.addMenu("Export")
        a_csv = QAction("Results as CSV…", self)
        a_csv.triggered.connect(self.export_results_csv)
        a_xlsx = QAction("Results as XLSX…", self)
        a_xlsx.triggered.connect(self.export_results_xlsx)
        a_md = QAction("Clips as Markdown…", self)
        a_md.triggered.connect(self.export_clips_markdown)
        a_ccsv = QAction("Clips as CSV…", self)
        a_ccsv.triggered.connect(self.export_clips_csv)
        m_export.addAction(a_csv)
        m_export.addAction(a_xlsx)
        m_export.addSeparator()
        m_export.addAction(a_md)
        m_export.addAction(a_ccsv)

        m_clips = mb.addMenu("Clips")
        a_show_clips = QAction("Show clips…", self)
        a_show_clips.triggered.connect(self.show_clips)
        a_clear_clips = QAction("Clear clips", self)
        a_clear_clips.triggered.connect(self.clear_clips)
        m_clips.addAction(a_show_clips)
        m_clips.addAction(a_clear_clips)

        m_tools = mb.addMenu("Tools")
        a_settings = QAction("Settings…", self)
        a_settings.triggered.connect(self.open_settings)
        a_clear_cache = QAction("Clear disk cache", self)
        a_clear_cache.triggered.connect(self.clear_cache)
        m_tools.addAction(a_settings)
        m_tools.addAction(a_clear_cache)

    # ----- status -----
    def _refresh_status(self, extra: str = ""):
        cs = "ON" if self.cfg.case_sensitive else "OFF"
        ww = "ON" if self.cfg.whole_word else "OFF"
        mode = self.cfg.search_mode
        view = self.cfg.results_view
        msg = f"Files: {len(self.files)} | Keywords: {len(self._parse_keywords())} | Mode: {mode} | Case {cs} | Whole {ww} | View: {view}"
        if extra:
            msg += f" — {extra}"
        self.status.setText(msg)

    # ----- parsing -----
    def _parse_keywords(self) -> List[str]:
        raw = (self.keywords_text or "").strip()
        if not raw:
            return []
        parts: List[str] = []
        for line in raw.splitlines():
            for p in line.split(","):
                k = p.strip()
                if k:
                    parts.append(k)
        # de-dup
        seen = set()
        out = []
        for k in parts:
            key = k if self.cfg.case_sensitive else k.lower()
            if key not in seen:
                seen.add(key)
                out.append(k)
        return out

    # ----- dialogs -----
    def manage_files(self):
        dlg = FileManagerDialog(self, self.files, start_dir=self.cfg.last_dir)
        if dlg.exec() == QDialog.Accepted:
            self.files = dlg.get_files()
            # update last dir
            if self.files:
                self.cfg.last_dir = str(Path(self.files[-1]).parent)
            save_config(self.paths.config_path, self.cfg)
            self._refresh_status()

    def clear_files(self):
        self.files = []
        self.hits = []
        self._view_rows = []
        self.table.setRowCount(0)
        self._refresh_status("Files cleared")

    def manage_keywords(self):
        dlg = KeywordManagerDialog(self, self.keywords_text, self.cfg)
        if dlg.exec() == QDialog.Accepted:
            self.keywords_text = dlg.get_text()
            dlg.apply_to_config(self.cfg)
            save_config(self.paths.config_path, self.cfg)
            self._refresh_status()

    def import_keywords(self):
        path, _ = QFileDialog.getOpenFileName(self, "Import keywords", self.cfg.last_dir or "", "Text/CSV (*.txt *.csv);;All files (*.*)")
        if not path:
            return
        p = Path(path)
        try:
            if p.suffix.lower() == ".csv":
                rows = []
                with open(p, "r", encoding="utf-8-sig", newline="") as f:
                    for r in csv.reader(f):
                        for cell in r:
                            cell = (cell or "").strip()
                            if cell:
                                rows.append(cell)
                self.keywords_text = "\n".join(rows)
            else:
                self.keywords_text = p.read_text(encoding="utf-8")
            self._refresh_status("Keywords imported")
        except Exception as e:
            QMessageBox.critical(self, "Import failed", str(e))

    def export_keywords(self):
        kws = self._parse_keywords()
        if not kws:
            QMessageBox.information(self, "Tip", "No keywords to export.")
            return
        out, _ = QFileDialog.getSaveFileName(self, "Export keywords", "keywords.txt", "Text (*.txt)")
        if not out:
            return
        Path(out).write_text("\n".join(kws), encoding="utf-8")
        QMessageBox.information(self, "Export", f"Saved:\n{out}")

    def export_prompt_template(self):
        tmpl = (
            "You are given academic documents (PDF/Word/PPT/Excel).\n"
            "Your task: output a clean keyword list to help locate key passages.\n\n"
            "Rules:\n"
            "- Output ONLY keywords, one per line.\n"
            "- No bullets, no numbering, no punctuation prefixes.\n"
            "- Avoid duplicates.\n"
            "- Prefer concrete phrases used in the text (2–6 words) when appropriate.\n\n"
            "Topic / question (fill in):\n"
            "<WRITE YOUR RESEARCH QUESTION HERE>\n\n"
            "Return 20–60 keywords:\n"
        )
        out, _ = QFileDialog.getSaveFileName(self, "Save prompt template", "keyword_prompt_template.txt", "Text (*.txt)")
        if not out:
            return
        Path(out).write_text(tmpl, encoding="utf-8")
        QMessageBox.information(self, "Export", f"Saved:\n{out}")

    def open_settings(self):
        dlg = SettingsDialog(self, self.cfg, self.paths)
        if dlg.exec() == QDialog.Accepted:
            dlg.apply()
            save_config(self.paths.config_path, self.cfg)
            self.cache.enable = self.cfg.enable_cache
            self._refresh_status("Settings saved")

    def clear_cache(self):
        self.cache.clear()
        QMessageBox.information(self, "Cache", "Disk cache cleared.")

    # ----- searching -----
    def run_search(self):
        if self.worker and self.worker.isRunning():
            QMessageBox.information(self, "Running", "Search is already running.")
            return
        if not self.files:
            QMessageBox.information(self, "Tip", "Add files first: Files → Manage files…")
            return
        keywords = self._parse_keywords()
        if not keywords:
            QMessageBox.information(self, "Tip", "Add keywords first: Keywords → Manage keywords…")
            return

        self.table.setRowCount(0)
        self.hits = []
        self._view_rows = []
        self.btn_run.setEnabled(False)
        self.btn_cancel.setEnabled(True)
        self.pbar.setRange(0, 1)
        self.pbar.setValue(0)
        self._refresh_status("Starting…")

        self.worker = SearchWorker(self.files, keywords, self.cfg, self.cache)
        self.worker.progress_text.connect(lambda t: self._refresh_status(t))
        self.worker.progress_range.connect(self.pbar.setRange)
        self.worker.progress_value.connect(self.pbar.setValue)
        self.worker.finished_ok.connect(self._on_search_ok)
        self.worker.finished_err.connect(self._on_search_err)
        self.worker.cancelled.connect(self._on_search_cancelled)
        self.worker.start()

    def cancel_search(self):
        if self.worker and self.worker.isRunning():
            self.worker.stop()
            self._refresh_status("Cancelling…")

    def _on_search_cancelled(self):
        self.btn_run.setEnabled(True)
        self.btn_cancel.setEnabled(False)
        self._refresh_status("Cancelled")

    def _on_search_err(self, msg: str):
        self.btn_run.setEnabled(True)
        self.btn_cancel.setEnabled(False)
        QMessageBox.critical(self, "Error", msg)
        self._refresh_status("Error")

    def _on_search_ok(self, hits: List[Hit]):
        self.btn_run.setEnabled(True)
        self.btn_cancel.setEnabled(False)
        self.hits = hits
        self.populate_table()
        if hits:
            self.table.selectRow(0)
            self._on_row_clicked(0, 0)

    # ----- views -----
    def set_results_view(self, which: str):
        which = "pages" if which == "pages" else "hits"
        self.cfg.results_view = which
        self.a_view_hits.setChecked(which == "hits")
        self.a_view_pages.setChecked(which == "pages")
        save_config(self.paths.config_path, self.cfg)
        self.populate_table()

    def populate_table(self):
        self.table.setRowCount(0)
        self._view_rows = []
        if not self.hits:
            self._refresh_status("No hits")
            return

        if self.cfg.results_view == "pages":
            groups: Dict[Tuple[str, str, Optional[int], Optional[int], Optional[int]], List[int]] = {}
            for i, h in enumerate(self.hits):
                key = (
                    h.file_path,
                    h.file_type,
                    h.page,
                    h.line_no if h.row_no is None else None,
                    h.row_no,
                )
                groups.setdefault(key, []).append(i)

            rows = list(groups.items())
            self.table.setRowCount(len(rows))
            for r, (key, idxs) in enumerate(rows):
                fpath, ftype, page, line, row_no = key
                rep = self.hits[idxs[0]]
                keywords = sorted({self.hits[i].keyword for i in idxs})
                detected = rep.detected_word
                snippet = rep.snippet

                self.table.setItem(r, 0, QTableWidgetItem("PAGE"))
                self.table.setItem(r, 1, QTableWidgetItem(ftype))
                self.table.setItem(r, 2, QTableWidgetItem(Path(fpath).name))
                self.table.setItem(r, 3, QTableWidgetItem(", ".join(keywords)[:120]))
                self.table.setItem(r, 4, QTableWidgetItem(detected))
                self.table.setItem(r, 5, QTableWidgetItem("" if page is None else str(page)))
                self.table.setItem(r, 6, QTableWidgetItem("" if line is None else str(line)))
                self.table.setItem(r, 7, QTableWidgetItem("" if row_no is None else str(row_no)))
                self.table.setItem(r, 8, QTableWidgetItem(f"[{len(idxs)} hits] {snippet}"))
                self._view_rows.append(idxs[0])
        else:
            self.table.setRowCount(len(self.hits))
            for i, h in enumerate(self.hits):
                self.table.setItem(i, 0, QTableWidgetItem("HIT"))
                self.table.setItem(i, 1, QTableWidgetItem(h.file_type))
                self.table.setItem(i, 2, QTableWidgetItem(Path(h.file_path).name))
                self.table.setItem(i, 3, QTableWidgetItem(h.keyword))
                self.table.setItem(i, 4, QTableWidgetItem(h.detected_word))
                self.table.setItem(i, 5, QTableWidgetItem("" if h.page is None else str(h.page)))
                self.table.setItem(i, 6, QTableWidgetItem("" if h.line_no is None else str(h.line_no)))
                self.table.setItem(i, 7, QTableWidgetItem("" if h.row_no is None else str(h.row_no)))
                self.table.setItem(i, 8, QTableWidgetItem(h.snippet))
                self._view_rows.append(i)

        self._refresh_status(f"Done. {len(self.hits)} hit(s).")
        self.setWindowTitle(f"Keyword Finder — {len(self.hits)} hit(s)")

    # ----- selection / navigation -----
    def _current_hit_index(self) -> Optional[int]:
        row = self.table.currentRow()
        if row < 0 or row >= len(self._view_rows):
            return None
        return self._view_rows[row]

    def _on_row_clicked(self, row: int, col: int):
        idx = self._current_hit_index()
        if idx is None:
            return
        if self.preview.isVisible():
            self.preview.show_hit(self.hits[idx])

    def select_next(self):
        if not self._view_rows:
            return
        r = self.table.currentRow()
        r = 0 if r < 0 else min(len(self._view_rows) - 1, r + 1)
        self.table.selectRow(r)
        self.table.scrollToItem(self.table.item(r, 0), QAbstractItemView.PositionAtCenter)
        if self.preview.isVisible():
            idx = self._current_hit_index()
            if idx is not None:
                self.preview.show_hit(self.hits[idx])

    def select_prev(self):
        if not self._view_rows:
            return
        r = self.table.currentRow()
        r = 0 if r < 0 else max(0, r - 1)
        self.table.selectRow(r)
        self.table.scrollToItem(self.table.item(r, 0), QAbstractItemView.PositionAtCenter)
        if self.preview.isVisible():
            idx = self._current_hit_index()
            if idx is not None:
                self.preview.show_hit(self.hits[idx])

    def toggle_preview(self):
        if self.preview.isVisible():
            self.preview.hide()
        else:
            self.preview.show()
            idx = self._current_hit_index()
            if idx is not None:
                self.preview.show_hit(self.hits[idx])

    # ----- external viewer -----
    def open_selected_in_viewer(self):
        idx = self._current_hit_index()
        if idx is None:
            QMessageBox.information(self, "Tip", "Select a result row first.")
            return
        hit = self.hits[idx]
        if hit.file_type != "PDF" or not hit.page:
            QMessageBox.information(self, "Tip", "Continue reading is available for PDF hits only.")
            return
        pdf_path = str(Path(hit.file_path).resolve())
        page = int(hit.page)

        mode = self.cfg.external_viewer
        if mode in ("auto", "sumatra") and _is_windows():
            exe = find_sumatra()
            if exe and mode in ("auto", "sumatra"):
                try:
                    subprocess.Popen([exe, "-reuse-instance", "-page", str(page), pdf_path], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
                    return
                except Exception as e:
                    logging.info("Sumatra open failed: %s", e)
            if mode == "sumatra":
                QMessageBox.information(self, "SumatraPDF not found", "SumatraPDF is not found. Falling back to default viewer (may not auto-jump to the page).")

        # fallback: default viewer
        url = QUrl.fromLocalFile(pdf_path)
        url.setFragment(f"page={page}")
        ok = QDesktopServices.openUrl(url)
        if not ok:
            ok2 = QDesktopServices.openUrl(QUrl.fromLocalFile(pdf_path))
            if not ok2:
                QMessageBox.warning(self, "Open failed", "Could not open the PDF in a viewer.\n\n" + pdf_path)

    # ----- clips -----
    def save_current_selection_as_clip(self):
        idx = self._current_hit_index()
        if idx is None:
            QMessageBox.information(self, "Tip", "Select a result row first.")
            return
        if not self.preview.isVisible():
            QMessageBox.information(self, "Tip", "Open the preview first: Preview → Show/Hide preview")
            return
        selected = self.preview.get_selected_text()
        if not selected:
            QMessageBox.information(self, "Tip", "Select some text in the Preview window first.")
            return
        hit = self.hits[idx]
        stamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.clips.append(
            Clip(
                created_at=stamp,
                file_path=hit.file_path,
                file_type=hit.file_type,
                page=hit.page,
                line_no=hit.line_no,
                row_no=hit.row_no,
                keyword=hit.keyword,
                detected_word=hit.detected_word,
                selected_text=selected,
            )
        )
        self._refresh_status(f"Saved clip. Total: {len(self.clips)}")
        if self.clips_dialog.isVisible():
            self.clips_dialog.refresh(self.clips)
            self.clips_dialog.select_last()

    def show_clips(self):
        self.clips_dialog.refresh(self.clips)
        self.clips_dialog.show()
        self.clips_dialog.raise_()
        self.clips_dialog.activateWindow()

    def clear_clips(self):
        self.clips = []
        self.clips_dialog.refresh(self.clips)
        self._refresh_status("Clips cleared")

    # ----- export -----
    def export_clips_markdown(self):
        if not self.clips:
            QMessageBox.information(self, "Tip", "No clips to export.")
            return
        default = f"quote_clips_{datetime.now().strftime('%Y%m%d_%H%M%S')}.md"
        out, _ = QFileDialog.getSaveFileName(self, "Save Markdown", default, "Markdown (*.md)")
        if not out:
            return
        lines = ["# Quote Clips\n", f"Exported: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n", "---\n"]
        for c in self.clips:
            where = ""
            if c.file_type == "PDF" and c.page is not None:
                where = f"Page {c.page}"
            elif c.row_no is not None:
                where = f"Row {c.row_no}"
            elif c.line_no is not None:
                where = f"Line {c.line_no}"
            lines.append(f"## {Path(c.file_path).name} — {where}\n")
            lines.append(f"- Saved: {c.created_at}")
            lines.append(f"- Keyword: {c.keyword}")
            if c.detected_word:
                lines.append(f"- Detected: {c.detected_word}")
            lines.append(f"- Source path: {c.file_path}\n")
            for ln in (c.selected_text or "").splitlines():
                lines.append(f"> {ln}")
            lines.append("\n")
        Path(out).write_text("\n".join(lines), encoding="utf-8")
        QMessageBox.information(self, "Export", f"Saved:\n{out}")

    def export_clips_csv(self):
        if not self.clips:
            QMessageBox.information(self, "Tip", "No clips to export.")
            return
        default = f"quote_clips_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        out, _ = QFileDialog.getSaveFileName(self, "Save CSV", default, "CSV (*.csv)")
        if not out:
            return
        with open(out, "w", newline="", encoding="utf-8-sig") as f:
            w = csv.writer(f)
            w.writerow(["created_at", "file", "file_type", "page", "line_no", "keyword", "detected_word", "selected_text"])
            for c in self.clips:
                w.writerow([c.created_at, c.file_path, c.file_type, c.page or "", c.line_no or "", c.keyword, c.detected_word, c.selected_text])
        QMessageBox.information(self, "Export", f"Saved:\n{out}")

    def export_clips_csv(self):
        if not self.clips:
            QMessageBox.information(self, "Tip", "No clips to export.")
            return
        default = f"quote_clips_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        out, _ = QFileDialog.getSaveFileName(self, "Save CSV", default, "CSV (*.csv)")
        if not out:
            return
        with open(out, "w", newline="", encoding="utf-8-sig") as f:
            w = csv.writer(f)
            w.writerow(["created_at", "file", "file_type", "page", "line_no", "row_no", "keyword", "detected_word", "selected_text"])
            for c in self.clips:
                w.writerow([
                    c.created_at,
                    c.file_path,
                    c.file_type,
                    c.page or "",
                    c.line_no or "",
                    c.row_no or "",
                    c.keyword,
                    c.detected_word,
                    c.selected_text,
                ])
        QMessageBox.information(self, "Export", f"Saved:\n{out}")

    def export_results_csv(self):
        if not self.hits:
            QMessageBox.information(self, "Tip", "No results to export.")
            return
        default = f"keyword_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        out, _ = QFileDialog.getSaveFileName(self, "Save CSV", default, "CSV (*.csv)")
        if not out:
            return
        with open(out, "w", newline="", encoding="utf-8-sig") as f:
            w = csv.writer(f)
            w.writerow(["file", "type", "keyword", "detected_word", "page", "line_no", "row_no", "x0", "y0", "x1", "y1", "snippet"])
            for h in self.hits:
                x0 = y0 = x1 = y1 = ""
                if h.rect:
                    x0, y0, x1, y1 = h.rect
                w.writerow([
                    h.file_path,
                    h.file_type,
                    h.keyword,
                    h.detected_word,
                    h.page or "",
                    h.line_no or "",
                    h.row_no or "",
                    x0,
                    y0,
                    x1,
                    y1,
                    h.snippet,
                ])
        QMessageBox.information(self, "Export", f"Saved:\n{out}")

    def export_results_xlsx(self):
        if not self.hits:
            QMessageBox.information(self, "Tip", "No results to export.")
            return
        default = f"keyword_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        out, _ = QFileDialog.getSaveFileName(self, "Save XLSX", default, "Excel (*.xlsx)")
        if not out:
            return
        wb = Workbook()
        ws = wb.active
        ws.title = "results"
        ws.append(["file", "type", "keyword", "detected_word", "page", "line_no", "row_no", "x0", "y0", "x1", "y1", "snippet"])
        for h in self.hits:
            x0 = y0 = x1 = y1 = None
            if h.rect:
                x0, y0, x1, y1 = h.rect
            ws.append([
                h.file_path,
                h.file_type,
                h.keyword,
                h.detected_word,
                h.page,
                h.line_no,
                h.row_no,
                x0,
                y0,
                x1,
                y1,
                h.snippet,
            ])
        ws.column_dimensions["A"].width = 46
        ws.column_dimensions["B"].width = 14
        ws.column_dimensions["C"].width = 22
        ws.column_dimensions["D"].width = 22
        ws.column_dimensions["E"].width = 8
        ws.column_dimensions["F"].width = 8
        ws.column_dimensions["G"].width = 8
        ws.column_dimensions["L"].width = 90
        wb.save(out)
        QMessageBox.information(self, "Export", f"Saved:\n{out}")

def main():
    portable = "--portable" in sys.argv
    paths = AppPaths(portable=portable)
    setup_logging(paths.logs_dir)
    cfg = load_config(paths.config_path)
    # ensure cache enable matches config
    app = QApplication([a for a in sys.argv if a != "--portable"])
    w = KeywordFinderWindow(paths, cfg)
    w.resize(1400, 860)
    w.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
